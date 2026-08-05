import base64
import io
import json
import os
import re
import urllib.error
import urllib.request
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


st.set_page_config(
    page_title="컨센서스 관리",
    page_icon="📝",
    layout="wide",
)

DATA_DIR = "data"
CONSENSUS_XLSX = os.path.join(DATA_DIR, "consensus.xlsx")
GITHUB_OWNER = "firestory119-hub"
GITHUB_REPO = "POR-Hunting"
GITHUB_PATH = "data/consensus.xlsx"
GITHUB_BRANCH = "main"
SHEET_NAME = "컨센서스입력"
HEADER_ROW = 2


def clean_ticker(value) -> str:
    text = str(value or "").strip().replace(".0", "")
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits.zfill(6) if digits else ""


@st.cache_data(show_spinner=False, ttl=300)
def load_consensus() -> pd.DataFrame:
    if not os.path.exists(CONSENSUS_XLSX):
        return pd.DataFrame()

    try:
        df = pd.read_excel(
            CONSENSUS_XLSX,
            sheet_name=SHEET_NAME,
            header=HEADER_ROW - 1,
            dtype={"종목코드": str},
            engine="openpyxl",
        )
    except Exception:
        return pd.DataFrame()

    if "종목코드" in df.columns:
        df["종목코드"] = df["종목코드"].map(clean_ticker)

    year_columns = [
        column
        for column in df.columns
        if re.fullmatch(r"\d{4}E?", str(column).strip())
    ]

    for column in year_columns:
        df[column] = pd.to_numeric(
            df[column],
            errors="coerce",
        )

    if "목표POR" in df.columns:
        df["목표POR"] = pd.to_numeric(
            df["목표POR"],
            errors="coerce",
        )

    if "업데이트일" in df.columns:
        df["업데이트일"] = pd.to_datetime(
            df["업데이트일"],
            errors="coerce",
        )

    return df


def sort_dataframe(
    df: pd.DataFrame,
    sort_option: str,
    selected_year: str | None,
) -> pd.DataFrame:
    result = df.copy()

    if sort_option == "종목명 가나다순" and "종목명" in result.columns:
        return result.sort_values(
            "종목명",
            ascending=True,
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    if sort_option == "종목명 역순" and "종목명" in result.columns:
        return result.sort_values(
            "종목명",
            ascending=False,
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    if sort_option == "최근 업데이트순" and "업데이트일" in result.columns:
        return result.sort_values(
            "업데이트일",
            ascending=False,
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    if sort_option == "오래된 업데이트순" and "업데이트일" in result.columns:
        return result.sort_values(
            "업데이트일",
            ascending=True,
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    if sort_option == "목표 POR 높은순" and "목표POR" in result.columns:
        return result.sort_values(
            "목표POR",
            ascending=False,
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    if sort_option == "목표 POR 낮은순" and "목표POR" in result.columns:
        return result.sort_values(
            "목표POR",
            ascending=True,
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    if (
        sort_option == "예상 영업이익 높은순"
        and selected_year
        and selected_year in result.columns
    ):
        return result.sort_values(
            selected_year,
            ascending=False,
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    if (
        sort_option == "예상 영업이익 낮은순"
        and selected_year
        and selected_year in result.columns
    ):
        return result.sort_values(
            selected_year,
            ascending=True,
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)

    return result.reset_index(drop=True)


def dataframe_to_excel_bytes(
    edited_df: pd.DataFrame,
) -> bytes:
    workbook = load_workbook(CONSENSUS_XLSX)

    if SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(
            f"엑셀에 '{SHEET_NAME}' 시트가 없습니다."
        )

    sheet = workbook[SHEET_NAME]

    header_map = {}

    for cell in sheet[HEADER_ROW]:
        if cell.value is not None:
            header_map[str(cell.value).strip()] = cell.column

    required_columns = [
        column
        for column in edited_df.columns
        if column in header_map
    ]

    last_existing_row = sheet.max_row
    first_data_row = HEADER_ROW + 1

    if last_existing_row >= first_data_row:
        for row_number in range(
            first_data_row,
            last_existing_row + 1,
        ):
            for column_name in required_columns:
                sheet.cell(
                    row=row_number,
                    column=header_map[column_name],
                ).value = None

    for row_offset, (_, row) in enumerate(
        edited_df.iterrows(),
        start=first_data_row,
    ):
        for column_name in required_columns:
            value = row.get(column_name)

            if pd.isna(value):
                value = None
            elif column_name == "종목코드":
                value = clean_ticker(value)
            elif column_name == "업데이트일":
                timestamp = pd.to_datetime(
                    value,
                    errors="coerce",
                )
                value = (
                    timestamp.to_pydatetime()
                    if pd.notna(timestamp)
                    else None
                )
            elif re.fullmatch(
                r"\d{4}E?",
                str(column_name).strip(),
            ):
                numeric_value = pd.to_numeric(
                    value,
                    errors="coerce",
                )
                value = (
                    float(numeric_value)
                    if pd.notna(numeric_value)
                    else None
                )
            elif column_name == "목표POR":
                numeric_value = pd.to_numeric(
                    value,
                    errors="coerce",
                )
                value = (
                    float(numeric_value)
                    if pd.notna(numeric_value)
                    else None
                )

            sheet.cell(
                row=row_offset,
                column=header_map[column_name],
            ).value = value

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_excel_to_github(
    excel_bytes: bytes,
) -> tuple[bool, str]:
    try:
        token = str(st.secrets["GITHUB_TOKEN"]).strip()
    except Exception:
        return (
            False,
            "Streamlit Secrets에 GITHUB_TOKEN이 없습니다.",
        )

    if not token:
        return (
            False,
            "Streamlit Secrets의 GITHUB_TOKEN이 비어 있습니다.",
        )

    api_url = (
        f"https://api.github.com/repos/"
        f"{GITHUB_OWNER}/{GITHUB_REPO}/contents/{GITHUB_PATH}"
    )

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "POR-Alpha-Consensus",
        "Content-Type": "application/json",
    }

    sha = None

    try:
        request = urllib.request.Request(
            api_url,
            method="GET",
            headers=headers,
        )

        with urllib.request.urlopen(
            request,
            timeout=20,
        ) as response:
            current_file = json.loads(
                response.read().decode("utf-8")
            )
            sha = current_file.get("sha")

    except urllib.error.HTTPError as error:
        if error.code != 404:
            detail = error.read().decode(
                "utf-8",
                errors="ignore",
            )[:400]
            return (
                False,
                f"GitHub 파일 확인 실패({error.code}): {detail}",
            )

    payload = {
        "message": "Update consensus data",
        "content": base64.b64encode(
            excel_bytes
        ).decode("ascii"),
        "branch": GITHUB_BRANCH,
    }

    if sha:
        payload["sha"] = sha

    request = urllib.request.Request(
        api_url,
        data=json.dumps(payload).encode("utf-8"),
        method="PUT",
        headers=headers,
    )

    try:
        with urllib.request.urlopen(
            request,
            timeout=40,
        ) as response:
            if response.status in (200, 201):
                return (
                    True,
                    "컨센서스를 GitHub에 저장했습니다.",
                )

            return (
                False,
                f"GitHub 저장 응답 코드: {response.status}",
            )

    except urllib.error.HTTPError as error:
        detail = error.read().decode(
            "utf-8",
            errors="ignore",
        )[:500]
        return (
            False,
            f"GitHub 저장 실패({error.code}): {detail}",
        )

    except Exception as error:
        return False, f"GitHub 저장 실패: {error}"


st.title("📝 컨센서스 관리")
st.caption(
    "종목별 연도 예상 영업이익을 수정하고 "
    "GitHub의 data/consensus.xlsx에 바로 저장합니다."
)

source_df = load_consensus()

if source_df.empty:
    st.error(
        "data/consensus.xlsx의 컨센서스입력 시트를 읽지 못했습니다."
    )
    st.stop()

year_columns = [
    str(column)
    for column in source_df.columns
    if re.fullmatch(r"\d{4}E?", str(column).strip())
]

registered_count = (
    source_df["종목명"].notna().sum()
    if "종목명" in source_df.columns
    else len(source_df)
)

input_count = (
    int(source_df[year_columns].notna().sum().sum())
    if year_columns
    else 0
)

latest_update = (
    source_df["업데이트일"].max()
    if "업데이트일" in source_df.columns
    else pd.NaT
)

metric1, metric2, metric3 = st.columns(3)
metric1.metric("등록 종목", f"{registered_count:,}개")
metric2.metric("영업이익 입력 수", f"{input_count:,}개")
metric3.metric(
    "최근 업데이트",
    (
        latest_update.strftime("%Y-%m-%d")
        if pd.notna(latest_update)
        else "-"
    ),
)

st.info(
    "새 종목은 표의 맨 아래 빈 행에 입력하세요. "
    "종목코드는 반드시 6자리로 입력합니다."
)

control1, control2, control3, control4 = st.columns(
    [2, 1.4, 1.2, 1.2]
)

with control1:
    search_text = st.text_input(
        "종목 검색",
        placeholder="종목명 또는 종목코드",
    ).strip()

with control2:
    sort_option = st.selectbox(
        "정렬 기준",
        [
            "종목명 가나다순",
            "종목명 역순",
            "최근 업데이트순",
            "오래된 업데이트순",
            "목표 POR 높은순",
            "목표 POR 낮은순",
            "예상 영업이익 높은순",
            "예상 영업이익 낮은순",
        ],
        index=0,
    )

with control3:
    selected_year = st.selectbox(
        "영업이익 정렬 연도",
        year_columns if year_columns else ["없음"],
        disabled=not bool(year_columns),
    )

with control4:
    filter_option = st.selectbox(
        "추가 필터",
        [
            "전체",
            "목표 POR 미입력",
            "업데이트일 미입력",
            "30일 이상 미갱신",
            "영업이익 입력 종목만",
        ],
    )

working_df = source_df.copy()

if search_text:
    name_mask = (
        working_df.get(
            "종목명",
            pd.Series("", index=working_df.index),
        )
        .astype(str)
        .str.contains(
            search_text,
            case=False,
            na=False,
        )
    )

    ticker_mask = (
        working_df.get(
            "종목코드",
            pd.Series("", index=working_df.index),
        )
        .astype(str)
        .str.contains(
            search_text,
            case=False,
            na=False,
        )
    )

    working_df = working_df[
        name_mask | ticker_mask
    ].copy()

if filter_option == "목표 POR 미입력":
    working_df = working_df[
        pd.to_numeric(
            working_df.get("목표POR"),
            errors="coerce",
        ).isna()
    ].copy()

elif filter_option == "업데이트일 미입력":
    working_df = working_df[
        pd.to_datetime(
            working_df.get("업데이트일"),
            errors="coerce",
        ).isna()
    ].copy()

elif filter_option == "30일 이상 미갱신":
    cutoff = (
        pd.Timestamp.today().normalize()
        - pd.Timedelta(days=30)
    )
    update_values = pd.to_datetime(
        working_df.get("업데이트일"),
        errors="coerce",
    )
    working_df = working_df[
        update_values.isna() | (update_values < cutoff)
    ].copy()

elif filter_option == "영업이익 입력 종목만":
    if year_columns:
        working_df = working_df[
            working_df[year_columns].notna().any(axis=1)
        ].copy()

working_df = sort_dataframe(
    working_df,
    sort_option,
    selected_year if year_columns else None,
)

display_columns = [
    column
    for column in [
        "종목명",
        "종목코드",
        *year_columns,
        "목표POR",
        "출처",
        "업데이트일",
        "비고",
    ]
    if column in working_df.columns
]

column_config = {
    "종목명": st.column_config.TextColumn(
        "종목명",
        required=True,
    ),
    "종목코드": st.column_config.TextColumn(
        "종목코드",
        help="6자리 숫자",
        validate=r"^\d{6}$",
    ),
}

for year in year_columns:
    column_config[year] = st.column_config.NumberColumn(
        f"{year} 영업이익(억)",
        format="%,.1f",
        step=1.0,
    )

if "목표POR" in display_columns:
    column_config["목표POR"] = (
        st.column_config.NumberColumn(
            "목표 POR",
            format="%.1f",
            step=0.5,
        )
    )

if "업데이트일" in display_columns:
    column_config["업데이트일"] = (
        st.column_config.DateColumn(
            "업데이트일",
            format="YYYY-MM-DD",
        )
    )

edited_visible = st.data_editor(
    working_df[display_columns],
    use_container_width=True,
    hide_index=True,
    num_rows="dynamic",
    column_config=column_config,
    key="consensus_main_editor_sorted",
)

st.caption(
    f"현재 화면: {len(working_df):,}개 종목 · "
    "정렬 상태에서 수정한 뒤 GitHub에 저장하세요."
)

button1, button2, button3 = st.columns(
    [1.1, 1.1, 2.2]
)

with button1:
    save_button = st.button(
        "💾 GitHub에 저장",
        type="primary",
        use_container_width=True,
    )

with button2:
    if st.button(
        "🔄 다시 읽기",
        use_container_width=True,
    ):
        load_consensus.clear()
        st.rerun()

with button3:
    try:
        current_excel_bytes = Path(
            CONSENSUS_XLSX
        ).read_bytes()
    except Exception:
        current_excel_bytes = b""

    st.download_button(
        "📥 현재 컨센서스 엑셀 다운로드",
        data=current_excel_bytes,
        file_name="consensus.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
        disabled=not bool(current_excel_bytes),
    )

if save_button:
    if search_text or filter_option != "전체":
        st.error(
            "검색 또는 추가 필터가 적용된 상태에서는 일부 종목만 보여 "
            "전체 저장이 차단됩니다. 검색어를 지우고 추가 필터를 "
            "'전체'로 바꾼 뒤 저장하세요."
        )
    else:
        try:
            excel_bytes = dataframe_to_excel_bytes(
                edited_visible
            )
            success, message = save_excel_to_github(
                excel_bytes
            )

            if success:
                st.success(message)
                st.cache_data.clear()
            else:
                st.error(message)

        except Exception as error:
            st.error(f"엑셀 생성 실패: {error}")

with st.expander("사용 방법"):
    st.markdown(
        """
1. 기본 화면은 **종목명 가나다순**입니다.
2. 상단에서 종목 검색, 정렬 기준, 연도, 추가 필터를 선택합니다.
3. 표의 값을 직접 수정하거나 맨 아래에 새 종목을 추가합니다.
4. 저장할 때는 검색어를 비우고 추가 필터를 `전체`로 변경합니다.
5. `GitHub에 저장`을 누르면 `data/consensus.xlsx`가 갱신됩니다.
        """
    )
